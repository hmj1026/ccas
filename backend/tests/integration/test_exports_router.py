"""Integration tests for /api/transactions/export (§8).

涵蓋：
- CSV streaming（含 unicode 商家名）
- xlsx export
- include_user_fields=true 時欄位齊全
- 日期 / 銀行 / 類別 filter
- 空資料路徑
"""

from __future__ import annotations

import io
from datetime import date, timedelta

import openpyxl
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from ccas.storage.models import BankConfig, Bill, Transaction
from tests.integration.conftest import auth_headers


async def _seed(
    session: AsyncSession,
    *,
    bank_code: str = "CTBC",
    billing_month: str = "2026-05",
    txns: list[tuple[int, str, str | None, str]] | None = None,
) -> Bill:
    """Seed bill + txns. txns: (amount, merchant, category, trans_date_iso)."""
    session.add(
        BankConfig(
            bank_code=bank_code,
            bank_name=f"{bank_code}-name",
            gmail_filter=f"from:{bank_code.lower()}",
        )
    )
    bill = Bill(
        bank_code=bank_code,
        billing_month=billing_month,
        total_amount=sum((t[0] for t in txns or []), 0),
        due_date=date.today() + timedelta(days=10),
        is_paid=False,
    )
    session.add(bill)
    await session.flush()
    for amount, merchant, cat, ds in txns or []:
        session.add(
            Transaction(
                bill_id=bill.id,
                trans_date=date.fromisoformat(ds),
                merchant=merchant,
                amount=amount,
                currency="TWD",
                category=cat,
                tags=["保留"],
                merchant_alias="",
            )
        )
    await session.commit()
    await session.refresh(bill)
    return bill


class TestCsvExport:
    async def test_empty(self, client: AsyncClient) -> None:
        resp = await client.get(
            "/api/transactions/export?format=csv", headers=auth_headers()
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("text/csv")
        body = resp.text.strip().splitlines()
        # 只有 header
        assert len(body) == 1
        assert "trans_date" in body[0]

    async def test_unicode_merchant(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        await _seed(
            db_session,
            txns=[
                (300, "星巴克咖啡", "餐飲", "2026-05-10"),
                (250, "全家便利店", "便利", "2026-05-11"),
            ],
        )
        resp = await client.get(
            "/api/transactions/export?format=csv", headers=auth_headers()
        )
        assert resp.status_code == 200
        body = resp.text
        assert "星巴克咖啡" in body
        assert "全家便利店" in body
        assert "300" in body

    async def test_filters_by_date_range(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        await _seed(
            db_session,
            txns=[
                (100, "M1", None, "2026-04-30"),
                (200, "M2", None, "2026-05-15"),
                (300, "M3", None, "2026-06-01"),
            ],
        )
        resp = await client.get(
            "/api/transactions/export?format=csv&start=2026-05-01&end=2026-05-31",
            headers=auth_headers(),
        )
        body = resp.text
        assert "M1" not in body
        assert "M2" in body
        assert "M3" not in body

    async def test_filters_by_bank(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        await _seed(
            db_session,
            bank_code="CTBC",
            txns=[(100, "C1", None, "2026-05-10")],
        )
        await _seed(
            db_session,
            bank_code="ESUN",
            txns=[(200, "E1", None, "2026-05-11")],
        )
        resp = await client.get(
            "/api/transactions/export?format=csv&bank=ESUN", headers=auth_headers()
        )
        body = resp.text
        assert "C1" not in body
        assert "E1" in body

    async def test_filters_by_category(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        await _seed(
            db_session,
            txns=[
                (100, "A1", "餐飲", "2026-05-10"),
                (200, "A2", "交通", "2026-05-11"),
            ],
        )
        resp = await client.get(
            "/api/transactions/export?format=csv&category=餐飲",
            headers=auth_headers(),
        )
        body = resp.text
        assert "A1" in body
        assert "A2" not in body

    async def test_include_user_fields_columns(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        await _seed(
            db_session,
            txns=[(100, "M1", "餐飲", "2026-05-10")],
        )
        resp = await client.get(
            "/api/transactions/export?format=csv&include_user_fields=true",
            headers=auth_headers(),
        )
        header = resp.text.splitlines()[0]
        for col in (
            "manual_category_override",
            "tags",
            "merchant_alias",
            "note",
        ):
            assert col in header


class TestXlsxExport:
    async def test_xlsx_returns_binary(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        await _seed(
            db_session,
            txns=[
                (100, "M1", "餐飲", "2026-05-10"),
                (200, "M2", "交通", "2026-05-11"),
            ],
        )
        resp = await client.get(
            "/api/transactions/export?format=xlsx", headers=auth_headers()
        )
        assert resp.status_code == 200
        ct = resp.headers["content-type"]
        assert "spreadsheetml" in ct or "octet-stream" in ct

        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True)
        ws = wb.active
        assert ws is not None
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 3  # header + 2 rows
        header_lower = [str(c or "").lower() for c in rows[0]]
        assert "trans_date" in header_lower
        assert "merchant" in header_lower
        # at least one txn row contains merchant we wrote
        merchant_idx = header_lower.index("merchant")
        merchants = {str(row[merchant_idx]) for row in rows[1:]}
        assert merchants == {"M1", "M2"}

    async def test_xlsx_include_user_fields(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        await _seed(
            db_session,
            txns=[(100, "M1", "餐飲", "2026-05-10")],
        )
        resp = await client.get(
            "/api/transactions/export?format=xlsx&include_user_fields=true",
            headers=auth_headers(),
        )
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True)
        ws = wb.active
        assert ws is not None
        rows = list(ws.iter_rows(values_only=True))
        header_lower = [str(c or "").lower() for c in rows[0]]
        for col in (
            "manual_category_override",
            "tags",
            "merchant_alias",
            "note",
        ):
            assert col in header_lower


class TestInvalidFormat:
    async def test_rejects_unknown_format(self, client: AsyncClient) -> None:
        resp = await client.get(
            "/api/transactions/export?format=pdf", headers=auth_headers()
        )
        assert resp.status_code == 422
